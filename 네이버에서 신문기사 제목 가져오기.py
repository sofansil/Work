import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

url = "https://search.naver.com/search.naver?where=nexearch&sm=top_hty&fbm=0&ie=utf8&query=%EC%95%84%EC%9D%B4%ED%8F%B017&ackey=wgrh0han"

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
}

response = requests.get(url, headers=headers)
response.encoding = 'utf-8'

soup = BeautifulSoup(response.text, 'html.parser')

# 뉴스 기사 제목 찾기 (headline1 클래스가 붙은 제목)
articles = soup.find_all('span', class_='sds-comps-text-ellipsis-1 sds-comps-text-type-headline1')

print("=== 신문기사 제목 ===")
titles = []

for article in articles:
    title_text = article.get_text(strip=True)
    # 제목이 비어있지 않으면 추가
    if title_text and len(titles) < 10:
        print(f"{len(titles) + 1}. {title_text}")
        titles.append(title_text)

# Excel 파일로 저장
wb = Workbook()
ws = wb.active
ws.title = "신문기사"

# 헤더 추가
ws['A1'] = "번호"
ws['B1'] = "제목"
ws['A1'].font = ws['B1'].font = Font(bold=True)

# 데이터 추가
for idx, title in enumerate(titles, 1):
    ws[f'A{idx+1}'] = idx
    ws[f'B{idx+1}'] = title

# 열 너비 조정
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 80

# 파일 저장
wb.save('results.xlsx')
print(f"\n✓ {len(titles)}개의 기사가 results.xlsx 파일로 저장되었습니다.")

